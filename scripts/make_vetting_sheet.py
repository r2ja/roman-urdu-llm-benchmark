#!/usr/bin/env python3
"""Build blind multi-annotator vetting workbooks (one .xlsx per annotator).

Each Pakistani reviewer fills their OWN copy independently (no peeking), then
merge_reviews.py combines them and computes agreement. See docs/VETTING.md.

Usage:
  python scripts/make_vetting_sheet.py --annotators 3        # review/annotator_1..3.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from rubench.metrics import MarkerLexicon, detect

REVIEW_DIR = ROOT / "review"
LEX = MarkerLexicon.load(ROOT / "datasets/markers/pakistani_markers.yaml")


def hindi_flag(text: str) -> str:
    """Pre-flag Roman-Hindi markers so reviewers can prioritise (e.g. 'Kripya')."""
    hits = detect(text or "", LEX).hindi
    return "HINDI? " + ", ".join(sorted(set(hits))) if hits else ""

INTENT_LABELS = ["balance_inquiry", "complaint", "refund_request", "order_status",
                 "technical_support", "account_security", "human_agent"]
SENT_LABELS = ["positive", "negative", "neutral"]

# task -> (path, kind, columns spec)
SHEETS = {
    "intent": (ROOT / "datasets/understanding/intent.jsonl", "classification", INTENT_LABELS),
    "sentiment": (ROOT / "datasets/understanding/sentiment.jsonl", "classification", SENT_LABELS),
    "translation": (ROOT / "datasets/understanding/translation_ur2en.jsonl", "translation", None),
    "generation": (ROOT / "datasets/generation/support_replies.jsonl", "generation", None),
}

HDR_FILL = PatternFill("solid", fgColor="E6E6E6")
HDR_FONT = Font(bold=True, color="000000")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(path: Path) -> list[dict]:
    out = []
    if not path.exists():
        return out
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("//"):
            out.append(json.loads(line))
    return out


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _dv(ws, options, col_letter, nrows):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    dv.error = "Pick from the list"; dv.prompt = "Choose one"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{nrows+1}")


def build_classification_sheet(ws, rows, labels):
    headers = ["id", "prompt (Roman Urdu)", "proposed_label", "auto_flag", "verdict", "corrected_label", "notes"]
    ws.append(headers)
    for r in rows:
        ws.append([r["id"], r["prompt"], r.get("label", ""), hindi_flag(r["prompt"]), "", "", ""])
    n = len(rows)
    _dv(ws, ["correct", "wrong", "unnatural", "drop"], "E", n)
    _dv(ws, labels, "F", n)
    widths = [12, 64, 18, 16, 12, 18, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _finalize(ws, n, len(headers), wrap_cols=[2])


def build_translation_sheet(ws, rows):
    headers = ["id", "Roman Urdu source", "proposed_english", "verdict", "corrected_english", "notes"]
    ws.append(headers)
    for r in rows:
        src = r["prompt"].replace("Translate to English: ", "")
        ws.append([r["id"], src, r.get("reference", ""), "", "", ""])
    n = len(rows)
    _dv(ws, ["good", "needs_fix", "drop"], "D", n)
    widths = [12, 50, 50, 12, 50, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _finalize(ws, n, len(headers), wrap_cols=[2, 3, 5])


def build_generation_sheet(ws, rows):
    headers = ["id", "scenario", "proposed_reply", "register", "auto_flag", "verdict", "corrected_reply", "notes"]
    ws.append(headers)
    for r in rows:
        ws.append([r["id"], r["prompt"], r.get("reference", ""), r.get("register", ""),
                   hindi_flag(r.get("reference", "")), "", "", ""])
    n = len(rows)
    _dv(ws, ["good", "needs_fix", "hindi_drift", "drop"], "F", n)
    widths = [12, 42, 52, 20, 16, 12, 52, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _finalize(ws, n, len(headers), wrap_cols=[2, 3, 7])


def _finalize(ws, nrows, ncols, wrap_cols):
    _style_header(ws, ncols)
    for row in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            if col in wrap_cols:
                cell.alignment = WRAP


def build_instructions(ws, annotator_id):
    lines = [
        ["Roman Urdu Benchmark — Human Vetting"],
        [f"Annotator copy #{annotator_id}. Fill this INDEPENDENTLY — do not discuss with other reviewers."],
        [""],
        ["You are a native Pakistani. Judge each item the way an ordinary Pakistani would."],
        [""],
        ["intent / sentiment tabs:"],
        ["  - verdict: 'correct' if the label fits; 'wrong' if a different label fits (then set corrected_label);"],
        ["    'unnatural' if the sentence is not natural Pakistani Roman Urdu; 'drop' to discard."],
        [""],
        ["translation tab:"],
        ["  - verdict: 'good' if the English faithfully conveys the Roman Urdu; 'needs_fix' (then fix corrected_english); 'drop'."],
        [""],
        ["generation tab:"],
        ["  - verdict: 'good' natural professional Pakistani reply; 'needs_fix' (fix corrected_reply);"],
        ["    'hindi_drift' if it uses Roman Hindi (dhanyavad, samasya, Desh...); 'drop'."],
        [""],
        ["Rules of thumb:"],
        ["  - English loanwords (account, balance, transfer, complaint) are FINE — that's authentic Pakistani Urdu."],
        ["  - Penalize Roman HINDI and stiff literary/'shudh' Urdu."],
        ["  - Flag anything a real Pakistani wouldn't actually type."],
    ]
    for ln in lines:
        ws.append(ln)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(bold=True, color="000000")
    ws.column_dimensions["A"].width = 110


def build_workbook(annotator_id: int) -> Workbook:
    wb = Workbook()
    info = wb.active; info.title = "Instructions"
    build_instructions(info, annotator_id)
    for task, (path, kind, labels) in SHEETS.items():
        rows = load(path)
        ws = wb.create_sheet(task)
        if kind == "classification":
            build_classification_sheet(ws, rows, labels)
        elif kind == "translation":
            build_translation_sheet(ws, rows)
        else:
            build_generation_sheet(ws, rows)
    return wb


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--annotators", type=int, default=3)
    args = ap.parse_args()
    REVIEW_DIR.mkdir(exist_ok=True)
    counts = {t: len(load(p)) for t, (p, _, _) in SHEETS.items()}
    for a in range(1, args.annotators + 1):
        wb = build_workbook(a)
        out = REVIEW_DIR / f"annotator_{a}.xlsx"
        wb.save(out)
        print(f"wrote {out}")
    print(f"\nItems per task: {counts}  (total {sum(counts.values())})")
    print(f"Hand each of the {args.annotators} files to a different Pakistani reviewer.")


if __name__ == "__main__":
    main()
