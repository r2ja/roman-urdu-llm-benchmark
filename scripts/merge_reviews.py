#!/usr/bin/env python3
"""Merge N filled annotator workbooks -> vetted gold set + agreement report.

Promotion rule (see docs/VETTING.md):
  - classification: gold label = majority corrected label; promote only if a strict
    majority agree on the same label. No majority -> discard.
  - translation/generation: promote if majority 'good' (or majority-supplied fix);
    majority 'drop'/'hindi_drift' -> discard.

Also reports inter-annotator agreement (Fleiss' kappa for classification,
% pairwise agreement for generation) — the credibility number for the dataset.

Usage:
  python scripts/merge_reviews.py review/ --out datasets/vetted/
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

SOURCE = {
    "intent": ROOT / "datasets/understanding/intent.jsonl",
    "sentiment": ROOT / "datasets/understanding/sentiment.jsonl",
    "translation": ROOT / "datasets/understanding/translation_ur2en.jsonl",
    "generation": ROOT / "datasets/generation/support_replies.jsonl",
}
CLASSIFICATION = {"intent", "sentiment"}


def load_jsonl(path):
    out = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("//"):
            r = json.loads(line); out[r["id"]] = r
    return out


def read_annotator(xlsx_path, task):
    """Return {id: {verdict, corrected_label/english/reply}} for one annotator+task."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if task not in wb.sheetnames:
        return {}
    ws = wb[task]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        rid = str(r[0])
        get = lambda col: (str(r[idx[col]]).strip() if col in idx and idx[col] < len(r) and r[idx[col]] not in (None, "") else "")
        out[rid] = {
            "verdict": get("verdict").lower(),
            "corrected_label": get("corrected_label").lower(),
            "corrected_english": get("corrected_english"),
            "corrected_reply": get("corrected_reply"),
        }
    return out


def fleiss_kappa(item_label_lists):
    """item_label_lists: list of lists of labels (one inner list per item, one label per annotator)."""
    cats = sorted({l for it in item_label_lists for l in it})
    if not cats or not item_label_lists:
        return None
    n = len(item_label_lists[0])
    if n < 2 or any(len(it) != n for it in item_label_lists):
        # ragged (some annotators skipped) — fall back to None
        return None
    N = len(item_label_lists)
    # P_i per item
    Pis = []
    cat_counts = Counter()
    for it in item_label_lists:
        c = Counter(it)
        Pis.append((sum(v * v for v in c.values()) - n) / (n * (n - 1)))
        cat_counts.update(it)
    Pbar = sum(Pis) / N
    pj = {k: v / (N * n) for k, v in cat_counts.items()}
    Pe = sum(v * v for v in pj.values())
    if Pe == 1:
        return 1.0
    return (Pbar - Pe) / (1 - Pe)


def majority(items):
    c = Counter(x for x in items if x)
    if not c:
        return None, 0
    top, cnt = c.most_common(1)[0]
    return top, cnt


def merge_classification(task, source, annos):
    ids = list(source.keys())
    N = len(annos)
    need = N // 2 + 1  # strict majority
    vetted, discarded = [], []
    label_lists = []
    for rid in ids:
        src = source[rid]
        proposed = src.get("label", "")
        final_labels = []
        for a in annos:
            v = a.get(rid, {})
            verdict = v.get("verdict", "")
            if verdict == "correct":
                final_labels.append(proposed)
            elif verdict == "wrong" and v.get("corrected_label"):
                final_labels.append(v["corrected_label"])
            elif verdict in ("unnatural", "drop"):
                final_labels.append("__DROP__")
            # blank verdict -> abstain (not counted)
        if len(final_labels) == N:
            label_lists.append(final_labels)
        lab, cnt = majority(final_labels)
        if lab and lab != "__DROP__" and cnt >= need:
            vetted.append({**src, "label": lab, "status": "vetted",
                           "n_agree": cnt, "n_annotators": len(final_labels)})
        else:
            discarded.append({"id": rid, "reason": "no_majority_or_drop",
                              "labels": final_labels})
    kappa = fleiss_kappa(label_lists) if label_lists else None
    return vetted, discarded, kappa


def merge_freeform(task, source, annos, corr_key, good_verdicts=("good",)):
    ids = list(source.keys())
    N = len(annos)
    need = N // 2 + 1
    vetted, discarded = [], []
    agree_count = 0; comparable = 0
    for rid in ids:
        src = source[rid]
        verdicts = []
        fixes = []
        for a in annos:
            v = a.get(rid, {})
            vd = v.get("verdict", "")
            if vd:
                verdicts.append(vd)
            if v.get(corr_key):
                fixes.append(v[corr_key])
        good = sum(1 for vd in verdicts if vd in good_verdicts)
        fixable = sum(1 for vd in verdicts if vd == "needs_fix")
        # pairwise agreement bookkeeping (good vs not-good)
        if len(verdicts) == N:
            comparable += 1
            binar = [1 if vd in good_verdicts else 0 for vd in verdicts]
            agree_count += 1 if len(set(binar)) == 1 else 0
        if good >= need:
            vetted.append({**src, "status": "vetted", "n_good": good})
        elif (good + fixable) >= need and fixes:
            # majority acceptable-with-fix; take the first supplied correction
            key = {"corrected_english": "reference", "corrected_reply": "reference"}[corr_key]
            vetted.append({**src, key: fixes[0], "status": "vetted",
                           "n_good": good, "fixed": True})
        else:
            discarded.append({"id": rid, "reason": "not_enough_good", "verdicts": verdicts})
    pct_agree = round(agree_count / comparable, 3) if comparable else None
    return vetted, discarded, pct_agree


def write_vetted(out_dir, task, rows):
    out_dir.mkdir(parents=True, exist_ok=True)
    dest = out_dir / (SOURCE[task].name)
    header = f"// VETTED gold for {task} — promoted by multi-annotator majority. See docs/VETTING.md"
    lines = [header] + [json.dumps({k: v for k, v in r.items()
                                    if k not in ("n_agree", "n_annotators", "n_good")},
                                   ensure_ascii=False) for r in rows]
    dest.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return dest


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("review_dir")
    ap.add_argument("--out", default="datasets/vetted")
    args = ap.parse_args()
    review_dir = Path(args.review_dir)
    files = sorted(review_dir.glob("annotator_*.xlsx"))
    if len(files) < 2:
        sys.exit(f"Need >=2 annotator files in {review_dir} (found {len(files)}).")
    out_dir = ROOT / args.out
    print(f"Merging {len(files)} annotators: {[f.name for f in files]}\n")

    report = {}
    for task, path in SOURCE.items():
        if not path.exists():
            continue
        source = load_jsonl(path)
        annos = [read_annotator(f, task) for f in files]
        if task in CLASSIFICATION:
            vetted, discarded, agr = merge_classification(task, source, annos)
            agr_name = "fleiss_kappa"
        else:
            corr = "corrected_english" if task == "translation" else "corrected_reply"
            vetted, discarded, agr = merge_freeform(task, source, annos, corr)
            agr_name = "pct_pairwise_agree"
        dest = write_vetted(out_dir, task, vetted)
        report[task] = {"candidates": len(source), "vetted": len(vetted),
                        "discarded": len(discarded), agr_name: agr}
        print(f"{task:12} candidates={len(source):3}  vetted={len(vetted):3}  "
              f"discarded={len(discarded):3}  {agr_name}={agr}")
        print(f"             -> {dest}")

    (out_dir / "agreement_report.json").write_text(json.dumps(report, indent=2))
    print(f"\nAgreement report -> {out_dir/'agreement_report.json'}")
    print("Target: kappa >= 0.6 (below that, tighten the task definition before trusting the data).")


if __name__ == "__main__":
    main()
